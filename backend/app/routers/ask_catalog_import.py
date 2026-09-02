"""Manager-facing bulk upload of the Marketplace's Ask catalog from a spreadsheet -
either replacing every existing Ask/Outcome/opt-in (Overwrite) or adding only the Asks
that aren't already in the Marketplace (Add), with a preview step before anything is
written. See `app/services/ask_catalog_import.py` for the parsing rules.
"""

import io

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.auth_context import Actor, get_current_actor, require_manager
from app.database import get_db
from app.models import (
    AiBreakdownRequest,
    Initiative,
    InitiativeEngineer,
    KbiCategory,
    KbiDetail,
    PlatformInitiativeCategory,
    PlatformInitiativeDetail,
    RecurringOpsCategory,
    RecurringOpsDetail,
    Task,
    TimeEntry,
    UpgradeUnit,
)
from app.models.enums import InitiativeStatus, InitiativeType
from app.schemas.ask_catalog_import import (
    AskCatalogCommitResult,
    AskCatalogImportMode,
    AskCatalogPreviewResponse,
    AskCatalogRowPreview,
)
from app.services.ask_catalog_import import (
    AskCatalogImportError,
    ParsedAskRow,
    TEMPLATE_HEADERS,
    parse_ask_catalog_workbook,
)

router = APIRouter(tags=["ask-catalog-import"])

# Manager Ask catalogs run to a few hundred rows at most - 10 MB is generous headroom.
_MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024

_CATEGORY_MODEL_BY_TYPE = {
    InitiativeType.RECURRING_OPS: RecurringOpsCategory,
    InitiativeType.PLATFORM: PlatformInitiativeCategory,
    InitiativeType.KBI: KbiCategory,
}
_DETAIL_MODEL_BY_TYPE = {
    InitiativeType.RECURRING_OPS: RecurringOpsDetail,
    InitiativeType.PLATFORM: PlatformInitiativeDetail,
    InitiativeType.KBI: KbiDetail,
}


async def _read_upload(file: UploadFile) -> bytes:
    contents = await file.read(_MAX_IMPORT_FILE_SIZE + 1)
    if len(contents) > _MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File is too large for an Ask catalog upload (max 10 MB)")
    return contents


def _to_row_preview(row: ParsedAskRow) -> AskCatalogRowPreview:
    return AskCatalogRowPreview(
        row_number=row.row_number,
        category=row.category,
        ask=row.ask,
        by_date=row.by_date,
        initiative_type=row.initiative_type,
        priority=row.priority,
        recurrence_type=row.recurrence_type,
        expected_delivery_date=row.expected_delivery_date,
        note=row.note,
        outcomes=row.outcomes,
        warnings=row.warnings,
    )


@router.post("/api/initiatives/import/ask-catalog/preview", response_model=AskCatalogPreviewResponse)
async def preview_ask_catalog_import(
    file: UploadFile = File(...),
    actor: Actor = Depends(get_current_actor),
):
    require_manager(actor)
    contents = await _read_upload(file)
    try:
        parsed = parse_ask_catalog_workbook(contents)
    except AskCatalogImportError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    valid = parsed.valid_rows
    return AskCatalogPreviewResponse(
        rows=[_to_row_preview(r) for r in parsed.rows],
        run_count=sum(1 for r in valid if r.initiative_type == InitiativeType.RECURRING_OPS),
        platform_count=sum(1 for r in valid if r.initiative_type == InitiativeType.PLATFORM),
        business_count=sum(1 for r in valid if r.initiative_type == InitiativeType.KBI),
        skipped_count=len(parsed.invalid_rows),
        total_outcomes=sum(len(r.outcomes) for r in valid),
    )


def _clear_ask_catalog(db: Session) -> None:
    """Wipes every Ask, Outcome, opt-in and category - but not Engineers - for a full
    Overwrite. Mirrors services/seed.py's _clear_all() minus the Engineer table.
    """
    for model in (
        TimeEntry,
        UpgradeUnit,
        Task,
        InitiativeEngineer,
        AiBreakdownRequest,
        KbiDetail,
        PlatformInitiativeDetail,
        RecurringOpsDetail,
        Initiative,
        KbiCategory,
        PlatformInitiativeCategory,
        RecurringOpsCategory,
    ):
        db.execute(delete(model))


def _get_or_create_category(db: Session, initiative_type: InitiativeType, name: str, cache: dict[str, object]):
    category = cache.get(name)
    if category is not None:
        return category, False
    model = _CATEGORY_MODEL_BY_TYPE[initiative_type]
    existing = db.query(model).filter(model.name == name).first()
    if existing is not None:
        cache[name] = existing
        return existing, False
    max_sort_order = db.query(model).count()
    created = model(name=name, sort_order=max_sort_order + 1)
    db.add(created)
    db.flush()
    cache[name] = created
    return created, True


def _create_initiative_with_outcomes(db: Session, row: ParsedAskRow, category) -> Initiative:
    assert row.initiative_type is not None
    initiative = Initiative(
        type=row.initiative_type,
        title=row.ask,
        description=row.note,
        expected_delivery_date=row.expected_delivery_date,
        priority=row.priority,
        status=InitiativeStatus.OPEN,
    )
    db.add(initiative)
    db.flush()

    detail_model = _DETAIL_MODEL_BY_TYPE[row.initiative_type]
    detail_kwargs = {"initiative_id": initiative.id, "category_id": category.id}
    if row.initiative_type == InitiativeType.RECURRING_OPS:
        detail_kwargs["recurrence_type"] = row.recurrence_type
        detail_kwargs["recurrence_interval"] = 1
    db.add(detail_model(**detail_kwargs))

    for order, outcome_title in enumerate(row.outcomes):
        db.add(
            Task(
                initiative_id=initiative.id,
                title=outcome_title,
                sequence_order=order,
                owner_engineer_id=None,
            )
        )

    return initiative


@router.post("/api/initiatives/import/ask-catalog/commit", response_model=AskCatalogCommitResult)
async def commit_ask_catalog_import(
    mode: AskCatalogImportMode = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_manager(actor)
    contents = await _read_upload(file)
    try:
        parsed = parse_ask_catalog_workbook(contents)
    except AskCatalogImportError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    valid_rows = parsed.valid_rows
    if not valid_rows:
        raise HTTPException(status_code=400, detail="No recognizable Ask rows found in this file")

    if mode == "overwrite":
        _clear_ask_catalog(db)

    category_cache: dict[str, object] = {}
    categories_created = 0
    skipped_existing_asks: list[str] = []
    counts = {InitiativeType.RECURRING_OPS: 0, InitiativeType.PLATFORM: 0, InitiativeType.KBI: 0}
    outcomes_created = 0

    for row in valid_rows:
        if mode == "add":
            exists = (
                db.query(Initiative.id)
                .filter(Initiative.type == row.initiative_type, Initiative.title.ilike(row.ask))
                .first()
            )
            if exists is not None:
                skipped_existing_asks.append(row.ask)
                continue

        category, created = _get_or_create_category(db, row.initiative_type, row.category, category_cache)
        if created:
            categories_created += 1
        _create_initiative_with_outcomes(db, row, category)
        counts[row.initiative_type] += 1
        outcomes_created += len(row.outcomes)

    db.commit()

    return AskCatalogCommitResult(
        mode=mode,
        run_count=counts[InitiativeType.RECURRING_OPS],
        platform_count=counts[InitiativeType.PLATFORM],
        business_count=counts[InitiativeType.KBI],
        outcomes_created=outcomes_created,
        categories_created=categories_created,
        skipped_existing_asks=skipped_existing_asks,
        skipped_unclassified_rows=len(parsed.invalid_rows),
    )


@router.get("/api/initiatives/import/ask-catalog/template")
def download_ask_catalog_template():
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Ask Catalog"
    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    ws.append(["Run Patching", "Monthly - Windows OS patching", "By end of month", "", "", ""])
    ws.append(
        [
            "Change Platform - UPS",
            "Refresh UPSes",
            "By end of year",
            "UPS-DC1-01",
            "UPS-DC1-02",
            "UPS-DC2-01",
        ]
    )
    ws.append(["Change Business - Amplify", "Amplify Go Live, incl all integrations and refreshes", "By Feb", "", "", ""])
    for col in "ABCDEF":
        ws.column_dimensions[col].width = 32

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ask-catalog-template.xlsx"'},
    )
