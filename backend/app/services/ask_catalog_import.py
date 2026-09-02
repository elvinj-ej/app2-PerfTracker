"""Parses a manager-uploaded Ask catalog workbook (a single sheet: Category | Ask | By
Date | Outcome 1 | Outcome 2 | ... | Outcome N) into structured rows ready to load into
the Marketplace, reusing the same "By Date" / priority rules as the one-time FY26-27
seed load (see `ask_parsing.py`).

The initiative type for each row is inferred from its Category text, mirroring the
convention already used in `app/data/fy2627_asks.py`: a category starting with "Run" is
Recurring Ops, "Change Platform" is a Platform Initiative, "Change Business" is a KBI.
Any other category is left unclassified and flagged - the row is shown in the preview
but not created.
"""

import io
from dataclasses import dataclass, field
from datetime import date

from openpyxl import load_workbook

from app.models.enums import InitiativeType, Priority, RecurrenceType
from app.services.ask_parsing import infer_priority, parse_change_delivery, parse_run_cadence

TEMPLATE_HEADERS = ["Category", "Ask", "By Date", "Outcome 1", "Outcome 2", "Outcome 3"]

_TYPE_PREFIXES: list[tuple[str, InitiativeType]] = [
    ("run", InitiativeType.RECURRING_OPS),
    ("change platform", InitiativeType.PLATFORM),
    ("change business", InitiativeType.KBI),
]


class AskCatalogImportError(ValueError):
    pass


def classify_category(category: str) -> InitiativeType | None:
    normalized = category.strip().lower()
    for prefix, initiative_type in _TYPE_PREFIXES:
        if normalized.startswith(prefix):
            return initiative_type
    return None


@dataclass
class ParsedAskRow:
    row_number: int
    category: str
    ask: str
    by_date: str | None
    outcomes: list[str]
    initiative_type: InitiativeType | None
    priority: Priority
    recurrence_type: RecurrenceType | None = None
    expected_delivery_date: date | None = None
    note: str | None = None
    warnings: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return self.initiative_type is not None


@dataclass
class ParsedAskCatalog:
    rows: list[ParsedAskRow]

    @property
    def valid_rows(self) -> list[ParsedAskRow]:
        return [r for r in self.rows if r.is_valid]

    @property
    def invalid_rows(self) -> list[ParsedAskRow]:
        return [r for r in self.rows if not r.is_valid]


def _find_outcome_columns(header_cells: list[str | None]) -> list[int]:
    return [
        idx
        for idx, value in enumerate(header_cells)
        if value is not None and str(value).strip().lower().startswith("outcome")
    ]


def _find_column(header_cells: list[str | None], name: str) -> int | None:
    """Matches a header starting with `name` (case-insensitive), so a manager's own
    variant - "Category - Initiative", "By Date (approx)" - is still recognized.
    """
    for idx, value in enumerate(header_cells):
        if value is not None and str(value).strip().lower().startswith(name):
            return idx
    return None


def parse_ask_catalog_workbook(contents: bytes) -> ParsedAskCatalog:
    try:
        workbook = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises various error types for a bad file
        raise AskCatalogImportError("Could not read this file as an Excel workbook (.xlsx)") from exc

    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration as exc:
        raise AskCatalogImportError("The workbook is empty") from exc

    category_col = _find_column(header, "category")
    ask_col = _find_column(header, "ask")
    by_date_col = _find_column(header, "by date")
    outcome_cols = _find_outcome_columns(header)

    if category_col is None or ask_col is None:
        raise AskCatalogImportError(
            "The first sheet must have 'Category' and 'Ask' column headers in row 1 "
            "(plus optional 'By Date' and one or more 'Outcome N' columns)"
        )

    seen_titles: dict[tuple[InitiativeType, str], int] = {}
    parsed_rows: list[ParsedAskRow] = []

    for offset, raw_row in enumerate(rows_iter, start=2):
        row = list(raw_row)

        def cell(col: int | None) -> str | None:
            if col is None or col >= len(row):
                return None
            value = row[col]
            if value is None:
                return None
            text = str(value).strip()
            return text or None

        category = cell(category_col) or ""
        ask = cell(ask_col) or ""
        by_date = cell(by_date_col)
        outcomes = [text for c in outcome_cols if (text := cell(c))]

        if not category and not ask and not outcomes:
            continue  # fully blank row - not a real data row

        if ask and category.strip().lower() == ask.strip().lower() and not by_date and not outcomes:
            continue  # a "RUN" / "CHANGE PLATFORM" / "CHANGE BUSINESS" section-header row, not a real Ask

        if not ask:
            parsed_rows.append(
                ParsedAskRow(
                    row_number=offset,
                    category=category,
                    ask=ask,
                    by_date=by_date,
                    outcomes=outcomes,
                    initiative_type=None,
                    priority=Priority.MEDIUM,
                    warnings=["Missing an Ask - row skipped"],
                )
            )
            continue

        initiative_type = classify_category(category)
        warnings: list[str] = []
        recurrence_type = None
        expected_delivery_date = None
        note = None

        if initiative_type is None:
            warnings.append(
                f"Category '{category}' doesn't start with Run / Change Platform / Change Business - row skipped"
            )
        elif initiative_type == InitiativeType.RECURRING_OPS:
            recurrence_type = parse_run_cadence(by_date)
        else:
            expected_delivery_date, note = parse_change_delivery(by_date)

        if initiative_type is not None:
            key = (initiative_type, ask.strip().lower())
            if key in seen_titles:
                warnings.append(f"Duplicate Ask title also seen at row {seen_titles[key]}")
            else:
                seen_titles[key] = offset

        parsed_rows.append(
            ParsedAskRow(
                row_number=offset,
                category=category,
                ask=ask,
                by_date=by_date,
                outcomes=outcomes,
                initiative_type=initiative_type,
                priority=infer_priority(category, ask),
                recurrence_type=recurrence_type,
                expected_delivery_date=expected_delivery_date,
                note=note,
                warnings=warnings,
            )
        )

    return ParsedAskCatalog(rows=parsed_rows)
