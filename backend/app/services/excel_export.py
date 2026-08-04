"""Builds the .xlsx workbook behind the Monthly Report's "Export to Excel" button.

Two sheets: a flat task-detail sheet (one row per task that had hours logged in the
selected month, across KBI/Platform/Recurring Ops) and a summary pivot of hours by
engineer x category, for quick reporting without needing to pivot in Excel first.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.enums import InitiativeType
from app.schemas.reporting import MonthlyInitiativeReport, MonthlyReport

CATEGORY_LABELS: dict[InitiativeType, str] = {
    InitiativeType.KBI: "Key Business Initiative",
    InitiativeType.PLATFORM: "Platform Initiative",
    InitiativeType.RECURRING_OPS: "Run Operations",
}

_HEADER_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _autosize_columns(ws: Worksheet, column_count: int) -> None:
    for col in range(1, column_count + 1):
        letter = get_column_letter(col)
        longest = max((len(str(cell.value)) for cell in ws[letter] if cell.value is not None), default=0)
        ws.column_dimensions[letter].width = max(12, longest + 3)


def build_monthly_report_workbook(report: MonthlyReport) -> bytes:
    wb = Workbook()

    detail_ws = wb.active
    detail_ws.title = "Task Detail"
    detail_headers = ["Category", "Initiative", "Task", "Stage", "Engineer", "Status", "Hours This Month"]
    _write_header(detail_ws, detail_headers)

    engineer_totals: dict[str, dict[str, float]] = {}
    groups: list[tuple[list[MonthlyInitiativeReport], InitiativeType]] = [
        (report.kbis, InitiativeType.KBI),
        (report.platform_initiatives, InitiativeType.PLATFORM),
        (report.recurring_ops, InitiativeType.RECURRING_OPS),
    ]

    for initiatives, itype in groups:
        category_label = CATEGORY_LABELS[itype]
        for initiative in initiatives:
            for task in initiative.tasks:
                if task.hours_this_month <= 0:
                    continue
                detail_ws.append(
                    [
                        category_label,
                        initiative.title,
                        task.title,
                        task.stage.value if task.stage else "",
                        task.owner_engineer_name,
                        task.status.value,
                        round(task.hours_this_month, 2),
                    ]
                )
                bucket = engineer_totals.setdefault(task.owner_engineer_name, {})
                bucket[category_label] = bucket.get(category_label, 0.0) + task.hours_this_month

    _autosize_columns(detail_ws, len(detail_headers))

    summary_ws = wb.create_sheet("Summary by Engineer")
    category_labels = list(CATEGORY_LABELS.values())
    summary_headers = ["Engineer", *category_labels, "Total"]
    _write_header(summary_ws, summary_headers)

    for engineer_name in sorted(engineer_totals):
        totals = engineer_totals[engineer_name]
        values = [round(totals.get(label, 0.0), 2) for label in category_labels]
        summary_ws.append([engineer_name, *values, round(sum(values), 2)])

    _autosize_columns(summary_ws, len(summary_headers))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
